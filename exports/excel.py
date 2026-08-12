"""
excel.py — Professional Industrial Digital Maturity Assessment Workbook.

Responsibilities
----------------
- Export assessment results to a professional Excel workbook.
- Present executive-level KPIs and maturity results.
- Export pillar, dimension, sub-dimension and indicator details.
- Export gap analysis and TPI prioritization.
- Export the priority matrix.
- Export the transformation roadmap.
- Export recommendations.
- Export assessment metadata.

This module does NOT perform calculations.
It only formats and presents already-computed results.

Workbook structure
------------------
01_Executive_Summary
02_DMI_Overview
03_Pillar_Assessment
04_Dimension_Assessment
05_Indicator_Details
06_Gap_Analysis
07_TPI_Prioritization
08_Priority_Matrix
09_Transformation_Roadmap
10_Recommendations
11_Assessment_Metadata
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Mapping, Optional

import pandas as pd

from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image as XLImage

from config import settings

from engines.assessment.scoring import ScoreResult
from engines.decision.gap import GapResult
from engines.decision.tpi import TPIResult
from engines.decision.priority import PriorityResult
from engines.decision.roadmap import RoadmapPhase

from utils.file_manager import ensure_directory, build_output_path


logger = settings.get_logger(__name__)


# ============================================================================
# BRAND / DESIGN SYSTEM
# ============================================================================

JESA_GREEN = "007A4D"
JESA_DARK = "173F35"
JESA_LIGHT = "E8F3EF"

ENSAM_BLUE = "0057A8"

WHITE = "FFFFFF"
BLACK = "000000"

GREY_100 = "F7F9F8"
GREY_200 = "E9EEEC"
GREY_400 = "B8C4BF"
GREY_600 = "5B6863"
GREY_800 = "25322E"

CRITICAL_RED = "C62828"
HIGH_ORANGE = "EF6C00"
MEDIUM_YELLOW = "F9A825"
LOW_BLUE = "1976D2"
VERY_LOW_GREY = "78909C"

SUCCESS_GREEN = "2E7D32"


# ============================================================================
# EXCEL EXPORTER
# ============================================================================


class ExcelExporter:
    """
    Generate a professional Industrial Digital Maturity Assessment workbook.

    Important design rule
    ---------------------
    Only real tabular data ranges are converted to Excel Tables.

    The Executive Summary contains merged presentation cells, therefore
    its title/metadata area is intentionally NOT converted to a Table.
    Only the KPI block is converted to a Table.
    """

    SHEETS = [
        "01_Executive_Summary",
        "02_DMI_Overview",
        "03_Pillar_Assessment",
        "04_Dimension_Assessment",
        "05_Indicator_Details",
        "06_Gap_Analysis",
        "07_TPI_Prioritization",
        "08_Priority_Matrix",
        "09_Transformation_Roadmap",
        "10_Recommendations",
        "11_Assessment_Metadata",
    ]

    # Sheets containing genuine tabular datasets.
    TABLE_SHEETS = {
        "02_DMI_Overview",
        "03_Pillar_Assessment",
        "04_Dimension_Assessment",
        "05_Indicator_Details",
        "06_Gap_Analysis",
        "07_TPI_Prioritization",
        "08_Priority_Matrix",
        "09_Transformation_Roadmap",
        "10_Recommendations",
        "11_Assessment_Metadata",
    }

    PRIORITY_COLORS = {
        "Critical": CRITICAL_RED,
        "High": HIGH_ORANGE,
        "Medium": MEDIUM_YELLOW,
        "Low": LOW_BLUE,
        "Very Low": VERY_LOW_GREY,
    }

    PRIORITY_RANK = {
        "Critical": 1,
        "High": 2,
        "Medium": 3,
        "Low": 4,
        "Very Low": 5,
    }

    PHASE_ORDER = {
        "Phase 1": 1,
        "Phase 2": 2,
        "Phase 3": 3,
        "Phase 4": 4,
    }

    def __init__(
        self,
        config: Optional[Mapping[str, Any]] = None,
    ) -> None:

        self.config = dict(config or {})

        self.decimal_precision = self.config.get(
            "decimal_precision",
            settings.SCORE_DECIMAL_PRECISION,
        )

        self.generated_at = datetime.now()

    # ========================================================================
    # PUBLIC API
    # ========================================================================

    def export_assessment_results(
        self,
        aggregation_results: Dict[str, Any],
        gap_results: Optional[List[GapResult]] = None,
        tpi_results: Optional[List[TPIResult]] = None,
        priority_results: Optional[List[PriorityResult]] = None,
        roadmap: Optional[List[RoadmapPhase]] = None,
        output_path: Optional[Path] = None,
        assessment_id: Optional[str] = None,
    ) -> Path:
        """
        Generate the complete Industrial Digital Maturity Assessment workbook.
        """

        if output_path is None:

            suffix = (
                f"_{assessment_id}"
                if assessment_id
                else ""
            )

            filename = (
                f"industrial_digital_maturity_assessment"
                f"{suffix}.xlsx"
            )

            output_path = build_output_path(filename)

        output_path = Path(output_path)

        ensure_directory(output_path.parent)

        # --------------------------------------------------------------------
        # CREATE WORKBOOK WITH PANDAS
        # --------------------------------------------------------------------

        with pd.ExcelWriter(
            output_path,
            engine="openpyxl",
        ) as writer:

            # --------------------------------------------------------------
            # 01 — Executive Summary
            # --------------------------------------------------------------

            self._write_executive_summary(
                writer,
                aggregation_results,
                gap_results,
                tpi_results,
                priority_results,
                roadmap,
                assessment_id,
            )

            # --------------------------------------------------------------
            # 02 — DMI Overview
            # --------------------------------------------------------------

            self._write_dmi_overview(
                writer,
                aggregation_results,
            )

            # --------------------------------------------------------------
            # 03 — Pillar Assessment
            # --------------------------------------------------------------

            self._write_pillar_assessment(
                writer,
                aggregation_results,
            )

            # --------------------------------------------------------------
            # 04 — Dimension Assessment
            # --------------------------------------------------------------

            self._write_dimension_assessment(
                writer,
                aggregation_results,
            )

            # --------------------------------------------------------------
            # 05 — Indicator Details
            # --------------------------------------------------------------

            self._write_indicator_details(
                writer,
                aggregation_results,
            )

            # --------------------------------------------------------------
            # 06 — Gap Analysis
            # --------------------------------------------------------------

            self._write_gap_analysis(
                writer,
                gap_results or [],
            )

            # --------------------------------------------------------------
            # 07 — TPI Prioritization
            # --------------------------------------------------------------

            self._write_tpi_prioritization(
                writer,
                tpi_results or [],
            )

            # --------------------------------------------------------------
            # 08 — Priority Matrix
            # --------------------------------------------------------------

            self._write_priority_matrix(
                writer,
                priority_results or [],
            )

            # --------------------------------------------------------------
            # 09 — Transformation Roadmap
            # --------------------------------------------------------------

            self._write_transformation_roadmap(
                writer,
                roadmap or [],
            )

            # --------------------------------------------------------------
            # 10 — Recommendations
            # --------------------------------------------------------------

            self._write_recommendations(
                writer,
                priority_results or [],
                roadmap or [],
            )

            # --------------------------------------------------------------
            # 11 — Assessment Metadata
            # --------------------------------------------------------------

            self._write_metadata(
                writer,
                aggregation_results,
                assessment_id,
            )

        # --------------------------------------------------------------------
        # POST-PROCESS WORKBOOK WITH OPENPYXL
        # --------------------------------------------------------------------

        self._format_workbook(
            output_path,
            aggregation_results,
        )

        logger.info(
            "Industrial Excel workbook generated: %s",
            output_path,
        )

        return output_path

    # ========================================================================
    # 01 — EXECUTIVE SUMMARY
    # ========================================================================

    def _write_executive_summary(
        self,
        writer: pd.ExcelWriter,
        results: Dict[str, Any],
        gap_results: Optional[List[GapResult]],
        tpi_results: Optional[List[TPIResult]],
        priority_results: Optional[List[PriorityResult]],
        roadmap: Optional[List[RoadmapPhase]],
        assessment_id: Optional[str],
    ) -> None:

        metadata = results.get(
            "metadata",
            {},
        )

        dmi = results.get("dmi")

        site_name = metadata.get(
            "site_name",
            "",
        )

        assessment_date = metadata.get(
            "assessment_date",
            "",
        )

        evaluator = metadata.get(
            "evaluator",
            "",
        )

        report_version = metadata.get(
            "report_version",
            "1.0",
        )

        rows = [
            {
                "Metric": "Digital Maturity Index",
                "Value": (
                    dmi.score
                    if dmi is not None
                    else metadata.get(
                        "dmi_score"
                    )
                ),
                "Unit": "%",
            },
            {
                "Metric": "Maturity Level",
                "Value": (
                    dmi.level_name
                    if dmi is not None
                    else metadata.get(
                        "dmi_level_name",
                        "",
                    )
                ),
                "Unit": "",
            },
            {
                "Metric": "Total Indicators",
                "Value": metadata.get(
                    "total_indicators",
                    0,
                ),
                "Unit": "Indicators",
            },
            {
                "Metric": "Total Dimensions",
                "Value": metadata.get(
                    "total_dimensions",
                    0,
                ),
                "Unit": "Dimensions",
            },
            {
                "Metric": "Total Pillars",
                "Value": metadata.get(
                    "total_pillars",
                    0,
                ),
                "Unit": "Pillars",
            },
            {
                "Metric": "Critical Priorities",
                "Value": self._count_priority(
                    priority_results,
                    "Critical",
                ),
                "Unit": "Actions",
            },
            {
                "Metric": "High Priorities",
                "Value": self._count_priority(
                    priority_results,
                    "High",
                ),
                "Unit": "Actions",
            },
            {
                "Metric": "Roadmap Actions",
                "Value": self._count_roadmap_items(
                    roadmap
                ),
                "Unit": "Actions",
            },
        ]

        df = pd.DataFrame(rows)

        df.to_excel(
            writer,
            sheet_name="01_Executive_Summary",
            index=False,
            startrow=8,
        )

        ws = writer.book["01_Executive_Summary"]

        # Main title
        ws["A1"] = (
            "INDUSTRIAL DIGITAL MATURITY "
            "ASSESSMENT REPORT"
        )

        ws["A2"] = (
            "Executive Assessment Workbook"
        )

        # Metadata
        ws["A4"] = "Assessment ID"
        ws["B4"] = assessment_id or metadata.get(
            "assessment_id",
            "",
        )

        ws["D4"] = "Site"
        ws["E4"] = site_name

        ws["A5"] = "Assessment Date"
        ws["B5"] = assessment_date

        ws["D5"] = "Evaluator"
        ws["E5"] = evaluator

        ws["A6"] = "Report Version"
        ws["B6"] = report_version

        ws["D6"] = "Generated"
        ws["E6"] = self.generated_at.strftime(
            "%Y-%m-%d %H:%M"
        )

        # KPI header
        ws["A9"] = "Executive KPI"
        ws["B9"] = "Value"
        ws["C9"] = "Unit"

        # IMPORTANT:
        # Pandas writes the dataframe starting at row 9 in Excel
        # because startrow=8 is zero-based.
        #
        # Therefore the KPI header is A9:C9.

        # Logo area
        self._insert_logos(ws)

    # ========================================================================
    # 02 — DMI OVERVIEW
    # ========================================================================

    def _write_dmi_overview(
        self,
        writer: pd.ExcelWriter,
        results: Dict[str, Any],
    ) -> None:

        dmi = results.get("dmi")

        rows = []

        if dmi is not None:

            rows.append(
                {
                    "Assessment Level": "Digital Maturity Index",
                    "ID": "DMI",
                    "Name": "Digital Maturity Index",
                    "Score": dmi.score,
                    "Maturity Level": dmi.level,
                    "Maturity Level Name": dmi.level_name,
                }
            )

        df = pd.DataFrame(rows)

        if df.empty:
            df = pd.DataFrame(
                columns=[
                    "Assessment Level",
                    "ID",
                    "Name",
                    "Score",
                    "Maturity Level",
                    "Maturity Level Name",
                ]
            )

        df.to_excel(
            writer,
            sheet_name="02_DMI_Overview",
            index=False,
        )

    # ========================================================================
    # 03 — PILLARS
    # ========================================================================

    def _write_pillar_assessment(
        self,
        writer: pd.ExcelWriter,
        results: Dict[str, Any],
    ) -> None:

        rows = []

        for _, result in results.get(
            "pillars",
            {},
        ).items():

            if isinstance(
                result,
                ScoreResult,
            ):

                rows.append(
                    {
                        "Pillar ID": result.entity_id,
                        "Pillar Name": result.entity_name,
                        "Score": result.score,
                        "Maturity Level": result.level,
                        "Maturity Level Name": result.level_name,
                        "Applicability": result.applicability,
                    }
                )

        df = pd.DataFrame(rows)

        if df.empty:
            df = pd.DataFrame(
                columns=[
                    "Pillar ID",
                    "Pillar Name",
                    "Score",
                    "Maturity Level",
                    "Maturity Level Name",
                    "Applicability",
                ]
            )

        df.to_excel(
            writer,
            sheet_name="03_Pillar_Assessment",
            index=False,
        )

    # ========================================================================
    # 04 — DIMENSIONS
    # ========================================================================

    def _write_dimension_assessment(
        self,
        writer: pd.ExcelWriter,
        results: Dict[str, Any],
    ) -> None:

        rows = []

        for _, result in results.get(
            "dimensions",
            {},
        ).items():

            if isinstance(
                result,
                ScoreResult,
            ):

                rows.append(
                    {
                        "Dimension ID": result.entity_id,
                        "Dimension Name": result.entity_name,
                        "Score": result.score,
                        "Maturity Level": result.level,
                        "Maturity Level Name": result.level_name,
                        "Applicability": result.applicability,
                        "Parent Pillar ID": result.parent_id,
                    }
                )

        df = pd.DataFrame(rows)

        if df.empty:
            df = pd.DataFrame(
                columns=[
                    "Dimension ID",
                    "Dimension Name",
                    "Score",
                    "Maturity Level",
                    "Maturity Level Name",
                    "Applicability",
                    "Parent Pillar ID",
                ]
            )

        df.to_excel(
            writer,
            sheet_name="04_Dimension_Assessment",
            index=False,
        )

    # ========================================================================
    # 05 — INDICATORS
    # ========================================================================

    def _write_indicator_details(
        self,
        writer: pd.ExcelWriter,
        results: Dict[str, Any],
    ) -> None:

        rows = []

        for _, result in results.get(
            "indicators",
            {},
        ).items():

            if isinstance(
                result,
                ScoreResult,
            ):

                rows.append(
                    {
                        "Indicator ID": result.entity_id,
                        "Indicator Name": result.entity_name,
                        "Score": result.score,
                        "Maturity Level": result.level,
                        "Maturity Level Name": result.level_name,
                        "Applicability": result.applicability,
                        "Parent ID": result.parent_id,
                    }
                )

        df = pd.DataFrame(rows)

        if df.empty:
            df = pd.DataFrame(
                columns=[
                    "Indicator ID",
                    "Indicator Name",
                    "Score",
                    "Maturity Level",
                    "Maturity Level Name",
                    "Applicability",
                    "Parent ID",
                ]
            )

        df.to_excel(
            writer,
            sheet_name="05_Indicator_Details",
            index=False,
        )

    # ========================================================================
    # 06 — GAP ANALYSIS
    # ========================================================================

    def _write_gap_analysis(
        self,
        writer: pd.ExcelWriter,
        gap_results: List[GapResult],
    ) -> None:

        rows = []

        for gap in gap_results:

            rows.append(
                {
                    "Entity ID": gap.entity_id,
                    "Entity Name": gap.entity_name,
                    "Entity Type": gap.entity_type,
                    "Current Score": gap.current_score,
                    "Target Score": gap.target_score,
                    "Gap": gap.gap,
                    "Gap (%)": gap.gap_percent,
                    "Priority": self._normalize_priority(
                        gap.priority
                    ),
                }
            )

        df = pd.DataFrame(rows)

        if df.empty:
            df = pd.DataFrame(
                columns=[
                    "Entity ID",
                    "Entity Name",
                    "Entity Type",
                    "Current Score",
                    "Target Score",
                    "Gap",
                    "Gap (%)",
                    "Priority",
                ]
            )

        df.to_excel(
            writer,
            sheet_name="06_Gap_Analysis",
            index=False,
        )

    # ========================================================================
    # 07 — TPI
    # ========================================================================

    def _write_tpi_prioritization(
        self,
        writer: pd.ExcelWriter,
        tpi_results: List[TPIResult],
    ) -> None:

        rows = []

        sorted_results = sorted(
            tpi_results,
            key=lambda x: (
                -float(x.tpi_score)
                if x.tpi_score is not None
                else 0,
                self.PRIORITY_RANK.get(
                    self._normalize_priority(
                        x.priority_category
                    ),
                    99,
                ),
            ),
        )

        for tpi in sorted_results:

            priority = self._normalize_priority(
                tpi.priority_category
            )

            rows.append(
                {
                    "Rank": len(rows) + 1,
                    "Dimension ID": tpi.dimension_id,
                    "Dimension Name": tpi.dimension_name,
                    "TPI Score": (
                        float(tpi.tpi_score) * 100
                        if tpi.tpi_score is not None
                        else None
                    ),
                    "Priority": priority,
                    "Gap": tpi.gap,
                    "Business Impact (%)": tpi.business_impact,
                    "Strategic Importance (%)": tpi.strategic_importance,
                    "Expected ROI (%)": tpi.expected_roi,
                    "Implementation Cost (MAD)": tpi.implementation_cost,
                    "Implementation Difficulty (person-days)": (
                        tpi.implementation_difficulty
                    ),
                }
            )

        df = pd.DataFrame(rows)

        if df.empty:
            df = pd.DataFrame(
                columns=[
                    "Rank",
                    "Dimension ID",
                    "Dimension Name",
                    "TPI Score",
                    "Priority",
                    "Gap",
                    "Business Impact (%)",
                    "Strategic Importance (%)",
                    "Expected ROI (%)",
                    "Implementation Cost (MAD)",
                    "Implementation Difficulty (person-days)",
                ]
            )

        df.to_excel(
            writer,
            sheet_name="07_TPI_Prioritization",
            index=False,
        )

    # ========================================================================
    # 08 — PRIORITY MATRIX
    # ========================================================================

    def _write_priority_matrix(
        self,
        writer: pd.ExcelWriter,
        priority_results: List[PriorityResult],
    ) -> None:

        rows = []

        sorted_results = sorted(
            priority_results,
            key=lambda x: (
                self.PRIORITY_RANK.get(
                    self._normalize_priority(
                        x.priority_category
                    ),
                    99,
                ),
                -(
                    x.tpi_score
                    if x.tpi_score is not None
                    else -1
                ),
                -(
                    x.gap
                    if x.gap is not None
                    else 0
                ),
            ),
        )

        for rank, result in enumerate(
            sorted_results,
            start=1,
        ):

            rows.append(
                {
                    "Rank": rank,
                    "Dimension ID": result.dimension_id,
                    "Dimension Name": result.dimension_name,
                    "Current Score": result.current_score,
                    "Target Score": result.target_score,
                    "Gap": result.gap,
                    "TPI Score": result.tpi_score,
                    "Priority": self._normalize_priority(
                        result.priority_category
                    ),
                    "Recommendations": len(
                        result.recommendations
                    ),
                }
            )

        df = pd.DataFrame(rows)

        if df.empty:
            df = pd.DataFrame(
                columns=[
                    "Rank",
                    "Dimension ID",
                    "Dimension Name",
                    "Current Score",
                    "Target Score",
                    "Gap",
                    "TPI Score",
                    "Priority",
                    "Recommendations",
                ]
            )

        df.to_excel(
            writer,
            sheet_name="08_Priority_Matrix",
            index=False,
        )

    # ========================================================================
    # 09 — ROADMAP
    # ========================================================================

    def _write_transformation_roadmap(
        self,
        writer: pd.ExcelWriter,
        roadmap: List[RoadmapPhase],
    ) -> None:

        rows = []

        sorted_phases = sorted(
            roadmap,
            key=lambda phase: self.PHASE_ORDER.get(
                phase.phase_name,
                99,
            ),
        )

        for phase in sorted_phases:

            sorted_items = sorted(
                phase.items,
                key=lambda item: (
                    -(
                        item.tpi_score
                        if item.tpi_score is not None
                        else 0
                    ),
                    -(
                        item.gap
                        if item.gap is not None
                        else 0
                    ),
                    item.title,
                ),
            )

            for item in sorted_items:

                rows.append(
                    {
                        "Phase": phase.phase_name,
                        "Horizon": phase.horizon,
                        "Recommendation ID": item.recommendation_id,
                        "Action": item.title,
                        "Dimension ID": item.dimension_id,
                        "Pillar ID": item.pillar_id,
                        "Priority": self._normalize_priority(
                            item.priority
                        ),
                        "TPI Score": item.tpi_score,
                        "Gap": item.gap,
                        "Effort": item.effort,
                        "Expected Impact": item.expected_impact,
                        "Prerequisites": "; ".join(
                            item.prerequisites or []
                        ),
                        "Dependencies": "; ".join(
                            item.dependencies or []
                        ),
                    }
                )

        df = pd.DataFrame(rows)

        if df.empty:
            df = pd.DataFrame(
                columns=[
                    "Phase",
                    "Horizon",
                    "Recommendation ID",
                    "Action",
                    "Dimension ID",
                    "Pillar ID",
                    "Priority",
                    "TPI Score",
                    "Gap",
                    "Effort",
                    "Expected Impact",
                    "Prerequisites",
                    "Dependencies",
                ]
            )

        df.to_excel(
            writer,
            sheet_name="09_Transformation_Roadmap",
            index=False,
        )

    # ========================================================================
    # 10 — RECOMMENDATIONS
    # ========================================================================

    def _write_recommendations(
        self,
        writer: pd.ExcelWriter,
        priority_results: List[PriorityResult],
        roadmap: List[RoadmapPhase],
    ) -> None:

        rows = []

        for priority in priority_results:

            normalized_priority = (
                self._normalize_priority(
                    priority.priority_category
                )
            )

            for recommendation in (
                priority.recommendations
            ):

                rows.append(
                    {
                        "Recommendation ID": (
                            recommendation.recommendation_id
                        ),
                        "Dimension ID": priority.dimension_id,
                        "Dimension Name": priority.dimension_name,
                        "Priority": normalized_priority,
                        "TPI Score": priority.tpi_score,
                        "Gap": priority.gap,
                        "Title": recommendation.title,
                        "Pillar ID": recommendation.pillar_id,
                        "Effort": recommendation.effort,
                        "Expected Impact": (
                            recommendation.expected_impact
                        ),
                        "Detailed Actions": "; ".join(
                            recommendation.detailed_actions
                            or []
                        ),
                        "Prerequisites": "; ".join(
                            recommendation.prerequisites
                            or []
                        ),
                        "Dependencies": "; ".join(
                            recommendation.dependencies
                            or []
                        ),
                        "Completion Evidence": "; ".join(
                            recommendation.completion_evidence
                            or []
                        ),
                        "Source Reference": (
                            recommendation.source_reference
                        ),
                    }
                )

        # If recommendations are absent from PriorityResult,
        # extract them directly from the roadmap.

        if not rows:

            for phase in roadmap:

                for item in phase.items:

                    if item.recommendation_id.startswith(
                        "NO-REC-"
                    ):
                        continue

                    rows.append(
                        {
                            "Recommendation ID": (
                                item.recommendation_id
                            ),
                            "Dimension ID": item.dimension_id,
                            "Dimension Name": "",
                            "Priority": self._normalize_priority(
                                item.priority
                            ),
                            "TPI Score": item.tpi_score,
                            "Gap": item.gap,
                            "Title": item.title,
                            "Pillar ID": item.pillar_id,
                            "Effort": item.effort,
                            "Expected Impact": (
                                item.expected_impact
                            ),
                            "Detailed Actions": "; ".join(
                                item.detailed_actions
                                or []
                            ),
                            "Prerequisites": "; ".join(
                                item.prerequisites
                                or []
                            ),
                            "Dependencies": "; ".join(
                                item.dependencies
                                or []
                            ),
                            "Completion Evidence": "; ".join(
                                item.completion_evidence
                                or []
                            ),
                            "Source Reference": (
                                item.source_reference
                            ),
                        }
                    )

        df = pd.DataFrame(rows)

        if df.empty:
            df = pd.DataFrame(
                columns=[
                    "Recommendation ID",
                    "Dimension ID",
                    "Dimension Name",
                    "Priority",
                    "TPI Score",
                    "Gap",
                    "Title",
                    "Pillar ID",
                    "Effort",
                    "Expected Impact",
                    "Detailed Actions",
                    "Prerequisites",
                    "Dependencies",
                    "Completion Evidence",
                    "Source Reference",
                ]
            )

        df.to_excel(
            writer,
            sheet_name="10_Recommendations",
            index=False,
        )

    # ========================================================================
    # 11 — METADATA
    # ========================================================================

    def _write_metadata(
        self,
        writer: pd.ExcelWriter,
        results: Dict[str, Any],
        assessment_id: Optional[str],
    ) -> None:

        metadata = results.get(
            "metadata",
            {},
        )

        rows = [
            {
                "Field": "Assessment ID",
                "Value": assessment_id
                or metadata.get(
                    "assessment_id",
                    "",
                ),
            },
            {
                "Field": "Site ID",
                "Value": metadata.get(
                    "site_id",
                    "",
                ),
            },
            {
                "Field": "Site Name",
                "Value": metadata.get(
                    "site_name",
                    "",
                ),
            },
            {
                "Field": "Assessment Date",
                "Value": metadata.get(
                    "assessment_date",
                    "",
                ),
            },
            {
                "Field": "Evaluator",
                "Value": metadata.get(
                    "evaluator",
                    "",
                ),
            },
            {
                "Field": "Report Version",
                "Value": metadata.get(
                    "report_version",
                    "1.0",
                ),
            },
            {
                "Field": "Generated At",
                "Value": self.generated_at.strftime(
                    "%Y-%m-%d %H:%M"
                ),
            },
            {
                "Field": "Total Indicators",
                "Value": metadata.get(
                    "total_indicators",
                    0,
                ),
            },
            {
                "Field": "Total Subdimensions",
                "Value": metadata.get(
                    "total_subdimensions",
                    0,
                ),
            },
            {
                "Field": "Total Dimensions",
                "Value": metadata.get(
                    "total_dimensions",
                    0,
                ),
            },
            {
                "Field": "Total Pillars",
                "Value": metadata.get(
                    "total_pillars",
                    0,
                ),
            },
            {
                "Field": "DMI Score",
                "Value": metadata.get(
                    "dmi_score",
                    None,
                ),
            },
            {
                "Field": "DMI Level",
                "Value": metadata.get(
                    "dmi_level",
                    "",
                ),
            },
            {
                "Field": "DMI Level Name",
                "Value": metadata.get(
                    "dmi_level_name",
                    "",
                ),
            },
        ]

        df = pd.DataFrame(rows)

        df.to_excel(
            writer,
            sheet_name="11_Assessment_Metadata",
            index=False,
        )

    # ========================================================================
    # WORKBOOK FORMATTING
    # ========================================================================

    def _format_workbook(
        self,
        output_path: Path,
        results: Dict[str, Any],
    ) -> None:

        wb = load_workbook(output_path)

        # --------------------------------------------------------------------
        # Guarantee exact sheet order
        # --------------------------------------------------------------------

        for sheet_name in self.SHEETS:

            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)

        wb._sheets = [
            wb[sheet_name]
            for sheet_name in self.SHEETS
        ]

        # --------------------------------------------------------------------
        # Base formatting
        # --------------------------------------------------------------------

        for ws in wb.worksheets:

            self._format_sheet(ws)

        # --------------------------------------------------------------------
        # Executive Summary
        # --------------------------------------------------------------------

        self._format_executive_summary(
            wb["01_Executive_Summary"]
        )

        # --------------------------------------------------------------------
        # KPI table
        # --------------------------------------------------------------------

        self._add_table(
            wb["01_Executive_Summary"],
            table_name="ExecutiveKPITable",
            start_row=9,
            end_row=wb[
                "01_Executive_Summary"
            ].max_row,
            start_col=1,
            end_col=3,
        )

        # --------------------------------------------------------------------
        # DMI
        # --------------------------------------------------------------------

        self._format_dmi(
            wb["02_DMI_Overview"]
        )

        # --------------------------------------------------------------------
        # Priority sheets
        # --------------------------------------------------------------------

        self._format_priority_sheet(
            wb["07_TPI_Prioritization"]
        )

        self._format_priority_sheet(
            wb["08_Priority_Matrix"]
        )

        # --------------------------------------------------------------------
        # Roadmap
        # --------------------------------------------------------------------

        self._format_roadmap(
            wb["09_Transformation_Roadmap"]
        )

        # --------------------------------------------------------------------
        # Recommendations
        # --------------------------------------------------------------------

        self._format_recommendations(
            wb["10_Recommendations"]
        )

        # --------------------------------------------------------------------
        # Metadata
        # --------------------------------------------------------------------

        self._format_metadata(
            wb["11_Assessment_Metadata"]
        )

        # --------------------------------------------------------------------
        # Final save
        # --------------------------------------------------------------------

        wb.save(output_path)

    # ========================================================================
    # BASE SHEET FORMAT
    # ========================================================================

    def _format_sheet(
        self,
        ws,
    ) -> None:

        ws.sheet_view.showGridLines = False

        max_row = ws.max_row
        max_col = ws.max_column

        # --------------------------------------------------------------------
        # Header formatting
        # --------------------------------------------------------------------

        if max_row >= 1 and max_col >= 1:

            self._style_header_row(
                ws,
                1,
            )

        # --------------------------------------------------------------------
        # IMPORTANT:
        #
        # We DO NOT create a worksheet-level AutoFilter.
        #
        # Excel Tables already provide their own filters.
        #
        # More importantly, the Executive Summary contains merged cells,
        # so a global table/filter starting at A1 would conflict with
        # the presentation layout and may cause Excel repair warnings.
        # --------------------------------------------------------------------

        if (
            ws.title in self.TABLE_SHEETS
            and max_row >= 2
            and max_col >= 1
        ):

            # Metadata is a key/value sheet and does not need a structured
            # Excel Table. This also keeps the workbook simpler.
            if ws.title != "11_Assessment_Metadata":

                self._add_table(
                    ws,
                    table_name=self._safe_table_name(
                        ws.title
                    ),
                )

        # --------------------------------------------------------------------
        # Freeze panes
        # --------------------------------------------------------------------

        if ws.title == "01_Executive_Summary":

            ws.freeze_panes = "A9"

        else:

            ws.freeze_panes = "A2"

        # --------------------------------------------------------------------
        # Column widths
        # --------------------------------------------------------------------

        for column_cells in ws.columns:

            if not column_cells:
                continue

            max_length = 0

            column_letter = (
                column_cells[0].column_letter
            )

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            ws.column_dimensions[
                column_letter
            ].width = min(
                max(max_length + 2, 12),
                55,
            )

        # --------------------------------------------------------------------
        # Wrap long textual fields
        # --------------------------------------------------------------------

        for row in ws.iter_rows():

            for cell in row:

                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

        # --------------------------------------------------------------------
        # Page setup
        # --------------------------------------------------------------------

        ws.page_setup.orientation = "landscape"

        ws.page_setup.paperSize = (
            ws.PAPERSIZE_A4
        )

        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.print_options.horizontalCentered = False

        ws.oddFooter.center.text = (
            "JESA DMAT | Industrial Digital "
            "Maturity Assessment"
        )

        ws.oddFooter.right.text = (
            "Page &[Page] of &[Pages]"
        )

        ws.oddHeader.right.text = (
            self.generated_at.strftime(
                "%Y-%m-%d"
            )
        )

        ws.sheet_properties.pageSetUpPr.autoPageBreaks = True

    # ========================================================================
    # EXECUTIVE SUMMARY FORMAT
    # ========================================================================

    def _format_executive_summary(
        self,
        ws,
    ) -> None:

        ws.freeze_panes = "A9"

        # --------------------------------------------------------------------
        # Merge ONLY presentation cells.
        #
        # These merged cells are intentionally outside the KPI table.
        # --------------------------------------------------------------------

        if "A1:H1" not in [
            str(rng)
            for rng in ws.merged_cells.ranges
        ]:

            ws.merge_cells("A1:H1")

        if "A2:H2" not in [
            str(rng)
            for rng in ws.merged_cells.ranges
        ]:

            ws.merge_cells("A2:H2")

        # --------------------------------------------------------------------
        # Main title
        # --------------------------------------------------------------------

        ws["A1"].font = Font(
            size=20,
            bold=True,
            color=WHITE,
        )

        ws["A1"].fill = PatternFill(
            "solid",
            fgColor=JESA_DARK,
        )

        ws["A1"].alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        # --------------------------------------------------------------------
        # Subtitle
        # --------------------------------------------------------------------

        ws["A2"].font = Font(
            size=11,
            italic=True,
            color=GREY_600,
        )

        ws["A2"].alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        # --------------------------------------------------------------------
        # Row heights
        # --------------------------------------------------------------------

        ws.row_dimensions[1].height = 34
        ws.row_dimensions[2].height = 22
        ws.row_dimensions[3].height = 55

        # --------------------------------------------------------------------
        # Metadata blocks
        # --------------------------------------------------------------------

        for cell in (
            "A4",
            "D4",
            "A5",
            "D5",
            "A6",
            "D6",
        ):

            ws[cell].font = Font(
                bold=True,
                color=JESA_DARK,
            )

            ws[cell].fill = PatternFill(
                "solid",
                fgColor=JESA_LIGHT,
            )

            ws[cell].alignment = Alignment(
                vertical="center",
                wrap_text=True,
            )

        # --------------------------------------------------------------------
        # KPI header
        #
        # The dataframe starts at Excel row 9.
        # --------------------------------------------------------------------

        for cell in (
            "A9",
            "B9",
            "C9",
        ):

            ws[cell].font = Font(
                bold=True,
                color=WHITE,
            )

            ws[cell].fill = PatternFill(
                "solid",
                fgColor=JESA_GREEN,
            )

            ws[cell].alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        # --------------------------------------------------------------------
        # KPI cards
        # --------------------------------------------------------------------

        for row in range(
            10,
            ws.max_row + 1,
        ):

            ws.cell(
                row,
                1,
            ).font = Font(
                bold=True,
                color=GREY_800,
            )

            ws.cell(
                row,
                2,
            ).font = Font(
                size=12,
                bold=True,
                color=JESA_DARK,
            )

            metric = ws.cell(
                row,
                1,
            ).value

            if metric == "Digital Maturity Index":

                ws.cell(
                    row,
                    2,
                ).number_format = "0.0"

            ws.cell(
                row,
                1,
            ).fill = PatternFill(
                "solid",
                fgColor=GREY_100,
            )

            ws.cell(
                row,
                2,
            ).fill = PatternFill(
                "solid",
                fgColor=JESA_LIGHT,
            )

        # --------------------------------------------------------------------
        # Executive summary print area
        # --------------------------------------------------------------------

        ws.print_area = (
            f"A1:H{max(ws.max_row, 18)}"
        )

    # ========================================================================
    # DMI FORMAT
    # ========================================================================

    def _format_dmi(
        self,
        ws,
    ) -> None:

        if ws.max_row >= 2:

            ws.freeze_panes = "A2"

            score_column = self._find_column(
                ws,
                "Score",
            )

            if score_column:

                ws.conditional_formatting.add(
                    f"{score_column}2:"
                    f"{score_column}{ws.max_row}",
                    ColorScaleRule(
                        start_type="min",
                        start_color=CRITICAL_RED,
                        mid_type="percentile",
                        mid_value=50,
                        mid_color=MEDIUM_YELLOW,
                        end_type="max",
                        end_color=SUCCESS_GREEN,
                    ),
                )

    # ========================================================================
    # PRIORITY FORMAT
    # ========================================================================

    def _format_priority_sheet(
        self,
        ws,
    ) -> None:

        priority_col = self._find_column(
            ws,
            "Priority",
        )

        if not priority_col:
            return

        for row in range(
            2,
            ws.max_row + 1,
        ):

            cell = ws[
                f"{priority_col}{row}"
            ]

            priority = (
                str(cell.value)
                if cell.value is not None
                else ""
            )

            color = self.PRIORITY_COLORS.get(
                priority
            )

            if color:

                cell.fill = PatternFill(
                    "solid",
                    fgColor=color,
                )

                cell.font = Font(
                    bold=True,
                    color=(
                        WHITE
                        if priority != "Medium"
                        else BLACK
                    ),
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                )

        tpi_col = self._find_column(
            ws,
            "TPI Score",
        )

        if tpi_col:

            ws.conditional_formatting.add(
                f"{tpi_col}2:"
                f"{tpi_col}{ws.max_row}",
                ColorScaleRule(
                    start_type="min",
                    start_color=GREY_200,
                    mid_type="percentile",
                    mid_value=50,
                    mid_color=MEDIUM_YELLOW,
                    end_type="max",
                    end_color=CRITICAL_RED,
                ),
            )

    # ========================================================================
    # ROADMAP FORMAT
    # ========================================================================

    def _format_roadmap(
        self,
        ws,
    ) -> None:

        priority_col = self._find_column(
            ws,
            "Priority",
        )

        phase_col = self._find_column(
            ws,
            "Phase",
        )

        if priority_col:

            for row in range(
                2,
                ws.max_row + 1,
            ):

                cell = ws[
                    f"{priority_col}{row}"
                ]

                priority = str(
                    cell.value or ""
                )

                color = self.PRIORITY_COLORS.get(
                    priority
                )

                if color:

                    cell.fill = PatternFill(
                        "solid",
                        fgColor=color,
                    )

                    cell.font = Font(
                        bold=True,
                        color=(
                            WHITE
                            if priority != "Medium"
                            else BLACK
                        ),
                    )

                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                    )

        if phase_col:

            for row in range(
                2,
                ws.max_row + 1,
            ):

                ws[
                    f"{phase_col}{row}"
                ].font = Font(
                    bold=True,
                    color=JESA_DARK,
                )

        # Wider action fields
        for col_name in (
            "Action",
            "Expected Impact",
            "Prerequisites",
            "Dependencies",
        ):

            col = self._find_column(
                ws,
                col_name,
            )

            if col:

                ws.column_dimensions[
                    col
                ].width = 40

    # ========================================================================
    # RECOMMENDATIONS FORMAT
    # ========================================================================

    def _format_recommendations(
        self,
        ws,
    ) -> None:

        self._format_priority_sheet(ws)

        for col_name in (
            "Title",
            "Detailed Actions",
            "Expected Impact",
            "Prerequisites",
            "Dependencies",
            "Completion Evidence",
        ):

            col = self._find_column(
                ws,
                col_name,
            )

            if col:

                ws.column_dimensions[
                    col
                ].width = 42

    # ========================================================================
    # METADATA FORMAT
    # ========================================================================

    def _format_metadata(
        self,
        ws,
    ) -> None:

        ws.freeze_panes = "A2"

        for row in range(
            2,
            ws.max_row + 1,
        ):

            ws.cell(
                row,
                1,
            ).font = Font(
                bold=True,
                color=JESA_DARK,
            )

            ws.cell(
                row,
                1,
            ).fill = PatternFill(
                "solid",
                fgColor=JESA_LIGHT,
            )

    # ========================================================================
    # LOGOS
    # ========================================================================

    def _insert_logos(
        self,
        ws,
    ) -> None:

        logo_dir = getattr(
            settings.frontend,
            "LOGO_DIR",
            Path(
                settings.backend.BASE_DIR
            )
            / "assets"
            / "logo",
        )

        logo_dir = Path(logo_dir)

        # JESA logo candidates
        jesa_candidates = [
            "jesa_logo.png",
            "JESA_logo.png",
            "jesa.png",
            "JESA.png",
        ]

        # ENSAM logo candidates
        ensam_candidates = [
            "ensam_logo.png",
            "ENSAM_logo.png",
            "ensam.png",
            "ENSAM.png",
        ]

        jesa_path = self._find_logo(
            logo_dir,
            jesa_candidates,
        )

        ensam_path = self._find_logo(
            logo_dir,
            ensam_candidates,
        )

        # --------------------------------------------------------------------
        # JESA
        # --------------------------------------------------------------------

        if jesa_path:

            try:

                img = XLImage(
                    str(jesa_path)
                )

                # Preserve original aspect ratio
                original_width = img.width
                original_height = img.height

                target_height = 55

                if original_height:

                    ratio = (
                        target_height
                        / original_height
                    )

                    img.height = target_height
                    img.width = int(
                        original_width * ratio
                    )

                ws.add_image(
                    img,
                    "G3",
                )

            except Exception as exc:

                logger.warning(
                    "Unable to insert JESA logo: %s",
                    exc,
                )

        # --------------------------------------------------------------------
        # ENSAM
        # --------------------------------------------------------------------

        if ensam_path:

            try:

                img = XLImage(
                    str(ensam_path)
                )

                # Preserve original aspect ratio
                original_width = img.width
                original_height = img.height

                target_height = 55

                if original_height:

                    ratio = (
                        target_height
                        / original_height
                    )

                    img.height = target_height
                    img.width = int(
                        original_width * ratio
                    )

                ws.add_image(
                    img,
                    "H3",
                )

            except Exception as exc:

                logger.warning(
                    "Unable to insert ENSAM logo: %s",
                    exc,
                )

    @staticmethod
    def _find_logo(
        directory: Path,
        candidates: List[str],
    ) -> Optional[Path]:

        if not directory.exists():
            return None

        for filename in candidates:

            path = directory / filename

            if path.exists():
                return path

        # Case-insensitive fallback
        candidate_names = {
            name.lower()
            for name in candidates
        }

        for path in directory.iterdir():

            if (
                path.is_file()
                and path.name.lower()
                in candidate_names
            ):

                return path

        return None

    # ========================================================================
    # TABLES
    # ========================================================================

    @staticmethod
    def _add_table(
        ws,
        table_name: str,
        start_row: int = 1,
        end_row: Optional[int] = None,
        start_col: int = 1,
        end_col: Optional[int] = None,
    ) -> None:
        """
        Add a structured Excel Table to a genuine tabular range.

        IMPORTANT:
        - Does not create worksheet-level AutoFilter.
        - Does not overlap merged cells.
        - Requires at least one header row and one data row.
        """

        if end_row is None:
            end_row = ws.max_row

        if end_col is None:
            end_col = ws.max_column

        # Need header + at least one data row.
        if end_row <= start_row:
            return

        if end_col < start_col:
            return

        start_letter = (
            ExcelExporter._column_letter(
                start_col
            )
        )

        end_letter = (
            ExcelExporter._column_letter(
                end_col
            )
        )

        ref = (
            f"{start_letter}{start_row}:"
            f"{end_letter}{end_row}"
        )

        # ------------------------------------------------------------
        # Prevent duplicate table names / tables.
        # ------------------------------------------------------------

        existing_table_names = {
            table.name
            for table in ws.parent.worksheets
            for table in table.tables.values()
        }

        if table_name in existing_table_names:
            return

        # ------------------------------------------------------------
        # Make sure the range does not intersect merged cells.
        # ------------------------------------------------------------

        for merged_range in ws.merged_cells.ranges:

            if (
                merged_range.min_row <= end_row
                and merged_range.max_row >= start_row
                and merged_range.min_col <= end_col
                and merged_range.max_col >= start_col
            ):
                logger.warning(
                    "Skipping Excel table '%s' because "
                    "its range %s intersects merged cells %s.",
                    table_name,
                    ref,
                    merged_range,
                )

                return

        table = Table(
            displayName=table_name,
            ref=ref,
        )

        style = TableStyleInfo(
            name="TableStyleMedium4",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )

        table.tableStyleInfo = style

        ws.add_table(table)

    # ========================================================================
    # HEADER
    # ========================================================================

    @staticmethod
    def _style_header_row(
        ws,
        row_number: int,
    ) -> None:

        for cell in ws[row_number]:

            cell.fill = PatternFill(
                "solid",
                fgColor=JESA_DARK,
            )

            cell.font = Font(
                bold=True,
                color=WHITE,
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

            cell.border = Border(
                bottom=Side(
                    style="thin",
                    color=WHITE,
                )
            )

        ws.row_dimensions[
            row_number
        ].height = 28

    # ========================================================================
    # HELPERS
    # ========================================================================

    @staticmethod
    def _normalize_priority(
        priority: Any,
    ) -> str:

        if priority is None:
            return ""

        mapping = {
            "critical": "Critical",
            "critique": "Critical",
            "high": "High",
            "haute": "High",
            "medium": "Medium",
            "moyenne": "Medium",
            "low": "Low",
            "faible": "Low",
            "very low": "Very Low",
            "très faible": "Very Low",
            "tres faible": "Very Low",
        }

        normalized = (
            str(priority)
            .strip()
            .lower()
        )

        return mapping.get(
            normalized,
            str(priority),
        )

    @staticmethod
    def _column_letter(
        column_number: int,
    ) -> str:

        result = ""

        while column_number:

            column_number, remainder = divmod(
                column_number - 1,
                26,
            )

            result = (
                chr(65 + remainder)
                + result
            )

        return result

    @staticmethod
    def _safe_table_name(
        name: str,
    ) -> str:

        cleaned = "".join(
            char
            if char.isalnum()
            else "_"
            for char in name
        )

        if not cleaned:
            cleaned = "AssessmentTable"

        if cleaned[0].isdigit():
            cleaned = f"T_{cleaned}"

        # Excel table names must not contain spaces
        # and must be reasonably short.
        return cleaned[:250]

    @staticmethod
    def _find_column(
        ws,
        header: str,
    ) -> Optional[str]:

        for cell in ws[1]:

            if str(
                cell.value
            ).strip() == header:

                return cell.column_letter

        return None

    @staticmethod
    def _count_priority(
        priority_results: Optional[
            List[PriorityResult]
        ],
        category: str,
    ) -> int:

        if not priority_results:
            return 0

        return sum(
            1
            for result in priority_results
            if ExcelExporter._normalize_priority(
                result.priority_category
            )
            == category
        )

    @staticmethod
    def _count_roadmap_items(
        roadmap: Optional[
            List[RoadmapPhase]
        ],
    ) -> int:

        if not roadmap:
            return 0

        return sum(
            len(phase.items)
            for phase in roadmap
        )


# ============================================================================
# PUBLIC UTILITY FUNCTIONS
# ============================================================================


def export_to_excel(
    results: Dict[str, Any],
    output_path: Optional[Path] = None,
    assessment_id: Optional[str] = None,
) -> Path:
    """
    Export aggregation results to a professional Excel workbook.
    """

    exporter = ExcelExporter()

    return exporter.export_assessment_results(
        aggregation_results=results,
        assessment_id=assessment_id,
        output_path=output_path,
    )


def export_full_analysis(
    aggregation_results: Dict[str, Any],
    gap_results: List[GapResult],
    tpi_results: List[TPIResult],
    priority_results: List[PriorityResult],
    roadmap: List[RoadmapPhase],
    output_path: Optional[Path] = None,
    assessment_id: Optional[str] = None,
) -> Path:
    """
    Export the complete assessment and decision analysis.
    """

    exporter = ExcelExporter()

    return exporter.export_assessment_results(
        aggregation_results=aggregation_results,
        gap_results=gap_results,
        tpi_results=tpi_results,
        priority_results=priority_results,
        roadmap=roadmap,
        output_path=output_path,
        assessment_id=assessment_id,
    )