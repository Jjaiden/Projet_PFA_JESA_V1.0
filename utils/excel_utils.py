"""
excel_utils.py — Lecture générique des classeurs Excel du projet.

Toutes les feuilles du référentiel JESA suivent le même format :
    ligne 1 : titre de la feuille (une seule cellule renseignée)
    ligne 2 : en-têtes de colonnes
    lignes 3+ : données

"""

from __future__ import annotations

import math
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.workbook import Workbook

TITLE_ROWS = 1  # nombre de lignes de titre avant l'en-tête, dans toutes les feuilles


def open_workbook(path: str | Path) -> Workbook:
    """Ouvre un classeur en lecture seule, valeurs calculées (pas de formules)."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Classeur introuvable : {path}")
    return openpyxl.load_workbook(path, data_only=True, read_only=True)


def sheet_to_records(wb: Workbook, sheet_name: str) -> list[dict[str, Any]]:
    """
    Convertit une feuille (titre + en-tête + données) en liste de dictionnaires
    {nom_de_colonne: valeur}, une entrée par ligne de données.

    Les lignes entièrement vides sont ignorées. Une ligne est considérée
    vide si sa première cellule est None (convention : la 1ère colonne est
    toujours un identifiant obligatoire dans nos feuilles).
    """
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"Feuille absente du classeur : {sheet_name}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) <= TITLE_ROWS:
        return []

    header = rows[TITLE_ROWS]
    data_rows = rows[TITLE_ROWS + 1 :]

    records: list[dict[str, Any]] = []
    for row in data_rows:
        if row is None or row[0] is None:
            continue
        record = {header[i]: row[i] for i in range(len(header)) if header[i] is not None}
        records.append(record)
    return records


def sheet_to_keyvalue(
    wb: Workbook,
    sheet_name: str,
    key_col: str = "Field_Name",
    value_col: str = "Field_Value",
) -> dict[str, Any]:
    """
    Convertit une feuille au format clé/valeur (ex. ASSESSMENT_METADATA) en
    dictionnaire simple {Field_Name: Field_Value}.
    """
    records = sheet_to_records(wb, sheet_name)
    return {r[key_col]: r.get(value_col) for r in records if key_col in r}


def group_records_by(records: list[dict[str, Any]], key: str) -> dict[Any, list[dict[str, Any]]]:
    """Regroupe une liste de dicts par la valeur d'une clé donnée (ex. Indicator_ID)."""
    grouped: dict[Any, list[dict[str, Any]]] = {}
    for r in records:
        grouped.setdefault(r.get(key), []).append(r)
    return grouped


def safe_int(value: Any, default: int | None = None) -> int | None:
    """Convertit uniquement un entier exact ; 3.0 est accepté, 3.5 est refusé."""
    if value is None or value == "" or isinstance(value, bool):
        return default

    try:
        numeric_value = float(value)
    except (TypeError, ValueError):
        return default

    if not math.isfinite(numeric_value) or not numeric_value.is_integer():
        return default

    return int(numeric_value)


def safe_float(value: Any, default: float | None = None) -> float | None:
    """Convertit en float fini en tolérant None et chaînes vides."""
    if value is None or value == "" or isinstance(value, bool):
        return default
    try:
        numeric_value = float(value)
    except (ValueError, TypeError):
        return default

    return numeric_value if math.isfinite(numeric_value) else default


def safe_str(value: Any, default: str | None = None) -> str | None:
    """Normalise une valeur texte : None/vide -> default, sinon strip()."""
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default
