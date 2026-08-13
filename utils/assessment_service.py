# utils/assessment_service.py
"""Streamlit-facing assessment workflow helpers.

This module adapts the existing backend engines to the frontend without
putting business calculations in the pages.
"""

from __future__ import annotations

import json
import math
import tempfile
import uuid
from dataclasses import replace
from pathlib import Path
from typing import Any, Mapping

import openpyxl
import pandas as pd

from config import settings
from config.constants import KnowledgeBaseSheets, RefSheets
from data.loader import load_assessment, load_referentiel
from data.models import AssessmentMetadata
from engines.assessment.aggregation import AggregationEngine
from engines.assessment.scoring import ScoreResult
from engines.assessment.validator import ValidationBlockingError, ensure_valid, validate_assessment
from engines.decision.gap import GapAnalysisEngine, GapResult
from engines.decision.priority import PriorityEngine, PriorityResult
from engines.decision.recommendation import RecommendationEngine, RecommendationResult
from engines.decision.roadmap import RoadmapEngine, RoadmapPhase
from engines.decision.tpi import TPIEngine, TPIResult


# ==============================================================================
# AJOUT DU LOGGER (CORRECTION DE L'ERREUR)
# ==============================================================================
logger = settings.get_logger(__name__)


REQUIRED_ASSESSMENT_SHEETS = {
    RefSheets.ASSESSMENT_METADATA,
    RefSheets.QUESTIONNAIRE_TEMPLATE,
}

REQUIRED_QUESTIONNAIRE_COLUMNS = {
    "Indicator_ID",
    "Pillar_ID",
    "Dimension_ID",
    "Subdimension_ID",
    "Selected_Score",
    "Applicability",
}


class AssessmentProcessingError(Exception):
    """User-facing assessment processing error."""


def save_uploaded_workbook(uploaded_file: Any) -> Path:
    """Persist a Streamlit UploadedFile to a temporary workbook path."""

    suffix = Path(uploaded_file.name).suffix.lower()
    if suffix != ".xlsx":
        raise AssessmentProcessingError("Only Excel files (.xlsx) are supported.")

    data = uploaded_file.getvalue()
    if not data:
        raise AssessmentProcessingError("The uploaded Excel file is empty.")

    temp_dir = Path(tempfile.gettempdir()) / "jesa_dmat_uploads"
    temp_dir.mkdir(parents=True, exist_ok=True)
    path = temp_dir / f"{uuid.uuid4().hex}{suffix}"
    path.write_bytes(data)
    return path


def validate_workbook_structure(path: Path) -> None:
    """Validate the uploaded assessment workbook before backend parsing."""

    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises several archive/parser errors
        raise AssessmentProcessingError(
            "The uploaded file is not a readable Excel workbook."
        ) from exc

    missing_sheets = REQUIRED_ASSESSMENT_SHEETS - set(workbook.sheetnames)
    if missing_sheets:
        raise AssessmentProcessingError(
            "The uploaded workbook is missing required sheet(s): "
            + ", ".join(sorted(missing_sheets))
            + "."
        )

    worksheet = workbook[RefSheets.QUESTIONNAIRE_TEMPLATE]
    rows = list(worksheet.iter_rows(values_only=True))
    if len(rows) <= 1:
        raise AssessmentProcessingError(
            "QUESTIONNAIRE_TEMPLATE must contain a header row and assessment rows."
        )

    columns = {str(value).strip() for value in rows[1] if value is not None}
    missing_columns = REQUIRED_QUESTIONNAIRE_COLUMNS - columns
    if missing_columns:
        raise AssessmentProcessingError(
            "QUESTIONNAIRE_TEMPLATE is missing required column(s): "
            + ", ".join(sorted(missing_columns))
            + "."
        )


def process_uploaded_assessment(
    path: Path,
    ui_metadata: Mapping[str, Any],
    source_filename: str,
) -> dict[str, Any]:
    """Run validation, scoring, gaps, recommendations, and dashboard adaptation."""

    validate_workbook_structure(path)

    # ★★★ CORRECTION ICI : charger le référentiel depuis le fichier uploadé ★★★
    referentiel = load_referentiel(file_path=path)

    assessment = load_assessment(path)

    assessment_id = f"ASM-{uuid.uuid4().hex[:12].upper()}"
    assessment.metadata = _merge_metadata(assessment.metadata, assessment_id, ui_metadata)

    validation_report = validate_assessment(assessment, referentiel)
    try:
        ensure_valid(validation_report)
    except ValidationBlockingError as exc:
        messages = [issue.message for issue in exc.report.errors[:8]]
        raise AssessmentProcessingError("Validation failed: " + " | ".join(messages)) from exc

    aggregation = AggregationEngine(referentiel).aggregate_scores(assessment)
    target_levels = _target_levels(referentiel)
    gaps = GapAnalysisEngine().calculate_dimension_gaps(
        dimension_scores=aggregation.get("dimensions", {}),
        target_levels=target_levels,
    )
    recommendations = _load_recommendations(gaps)
    summary = _build_summary(aggregation, gaps)

    backend_results = {
        "assessment_id": assessment_id,
        "metadata": {
            **aggregation.get("metadata", {}),
            "assessment_name": ui_metadata.get("assessment_name") or assessment_id,
            "company": ui_metadata.get("company"),
            "assessor_name": ui_metadata.get("assessor_name"),
            "assessor_role": ui_metadata.get("assessor_role"),
            "contact_email": ui_metadata.get("contact_email"),
            "source_filename": source_filename,
        },
        "aggregation": aggregation,
        "gaps": gaps,
        "recommendations": recommendations,
        "tpi": [],
        "priorities": [],
        "roadmap": [],
        "summary": summary,
    }

    return {
        "assessment_id": assessment_id,
        "backend_results": backend_results,
        "dashboard_data": build_dashboard_data(backend_results),
        "roadmap_results": None,
        "serialized_results": serialize_backend_results(backend_results),
        "validation_warnings": [issue.message for issue in validation_report.warnings],
    }


def run_decision_analysis(
    backend_results: dict[str, Any],
    decision_inputs: Mapping[str, Mapping[str, Any]],
) -> dict[str, Any]:
    """Run TPI, priority, and roadmap from page-provided decision inputs."""

    gaps = backend_results.get("gaps") or []
    recommendations = backend_results.get("recommendations") or {}
    tpi = TPIEngine().calculate_tpi(gaps, decision_inputs)
    priorities = PriorityEngine().build_priority_analysis(gaps, tpi, recommendations)
    roadmap = RoadmapEngine().build_roadmap(tpi, recommendations)

    backend_results["tpi"] = tpi
    backend_results["priorities"] = priorities
    backend_results["roadmap"] = roadmap
    backend_results["summary"] = _build_summary(backend_results["aggregation"], gaps, tpi, priorities)

    roadmap_results = build_roadmap_view_data(backend_results)
    return {
        "backend_results": backend_results,
        "roadmap_results": roadmap_results,
        "serialized_results": serialize_backend_results(backend_results),
    }


def build_dashboard_data(backend_results: Mapping[str, Any]) -> dict[str, Any]:
    """Adapt backend aggregation and gap outputs to the existing Dashboard UI."""

    aggregation = backend_results.get("aggregation", {})
    dmi = aggregation.get("dmi")
    gaps_by_dimension = {gap.entity_id: gap for gap in backend_results.get("gaps", [])}

    pillars = [
        {"name": result.entity_name, "score": round(float(result.score) * 20, 1)}
        for result in aggregation.get("pillars", {}).values()
        if isinstance(result, ScoreResult)
    ]

    dimensions = []
    for result in aggregation.get("dimensions", {}).values():
        if not isinstance(result, ScoreResult):
            continue
        gap = gaps_by_dimension.get(result.entity_id)
        target = gap.target_score if gap else result.score
        dimensions.append(
            {
                "id": result.entity_id,
                "name": result.entity_name,
                "current": round(float(result.score) * 20, 1),
                "target": round(float(target) * 20, 1),
                "gap": round((float(result.score) - float(target)) * 20, 1),
                "gap_raw": round(float(gap.gap), 2) if gap else 0.0,
            }
        )

    target_dmi = _target_dmi_percent(backend_results.get("gaps", []), aggregation)
    dmi_score = float(dmi.score) if dmi else 0.0
    largest_gap = min(dimensions, key=lambda item: item["gap"], default=None)
    strongest = max(pillars, key=lambda item: item["score"], default=None)

    return {
        "dmi": dmi_score,
        "maturity_level": dmi.level_name if dmi else "N/A",
        "target_dmi": target_dmi,
        "dmi_gap": round(dmi_score - target_dmi, 1),
        "pillars": pillars,
        "dimensions": sorted(dimensions, key=lambda item: item["gap"]),
        "insights": _dashboard_insights(largest_gap, strongest),
    }


def build_decision_rows(backend_results: Mapping[str, Any]) -> list[dict[str, Any]]:
    """Return dimensions with positive gaps for the Decision Analysis table."""

    rows = []
    for gap in backend_results.get("gaps", []):
        if gap.gap <= 0:
            continue
        rows.append(
            {
                "dimension_id": gap.entity_id,
                "dimension": gap.entity_name,
                "current_score": round(gap.current_score, 2),
                "target_score": round(gap.target_score, 2),
                "gap": round(gap.gap, 2),
            }
        )
    return rows


def _priority_from_tpi(
    tpi_score: float,
) -> str:
    """
    Derive the displayed priority strictly from the TPI score.

    Internal TPI:
        0.80 - 1.00 -> Critical
        0.60 - 0.7999 -> High
        0.40 - 0.5999 -> Medium
        0.20 - 0.3999 -> Low
        0.00 - 0.1999 -> Very Low

    The UI displays TPI as a percentage.
    """

    try:
        score = float(tpi_score)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Invalid TPI score: {tpi_score!r}") from exc

    if not math.isfinite(score):
        raise ValueError(f"TPI score must be finite: {score!r}")

    # Accept both:
    # 0.703
    # and
    # 70.3
    if score > 1.0:
        if score <= 100.0:
            score /= 100.0
        else:
            raise ValueError(f"TPI score outside valid range: {score}")

    score = max(0.0, min(1.0, score))

    if score >= 0.80:
        return "Critical"
    if score >= 0.60:
        return "High"
    if score >= 0.40:
        return "Medium"
    if score >= 0.20:
        return "Low"
    return "Very Low"


def build_roadmap_view_data(
    backend_results: Mapping[str, Any],
) -> dict[str, Any]:
    """
    Build the Roadmap page data.

    IMPORTANT:
    - TPI is the single source of truth for ranking.
    - Priority category is derived directly from TPI.
    - Actions are globally sorted by TPI descending.
    - TPI is displayed as a percentage.
    """

    aggregation = backend_results.get("aggregation") or {}
    dmi = aggregation.get("dmi")

    actions: list[dict[str, Any]] = []
    roadmap = backend_results.get("roadmap") or []

    for phase in roadmap:
        phase_items = getattr(phase, "items", None) or []
        for item in phase_items:
            raw_tpi = getattr(item, "tpi_score", 0)
            try:
                tpi_internal = float(raw_tpi or 0)
            except (TypeError, ValueError):
                tpi_internal = 0.0

            # Defensive normalization:
            # internal TPI must be [0,1]
            if tpi_internal > 1.0:
                if tpi_internal <= 100.0:
                    tpi_internal /= 100.0
                else:
                    tpi_internal = 1.0

            tpi_internal = max(0.0, min(1.0, tpi_internal))
            tpi_percent = round(tpi_internal * 100, 1)
            priority = _priority_from_tpi(tpi_internal)

            detailed_actions = getattr(item, "detailed_actions", None) or []

            actions.append(
                {
                    "rank": 0,
                    "title": getattr(item, "title", ""),
                    "dimension": getattr(item, "dimension_id", ""),
                    "tpi": tpi_percent,
                    "priority_level": priority,
                    "phase": f"{getattr(phase, 'phase_name', '')} ({getattr(phase, 'horizon', '')})",
                    "description": getattr(item, "expected_impact", ""),
                    "objective": getattr(item, "title", ""),
                    "expected_benefit": getattr(item, "expected_impact", ""),
                    "implementation_note": "; ".join(detailed_actions[:3]),
                }
            )

    # Global TPI sort
    actions.sort(key=lambda action: float(action["tpi"]), reverse=True)

    # Global rank
    for rank, action in enumerate(actions, start=1):
        action["rank"] = rank

    # Summary
    return {
        "summary": {
            "dmi": float(dmi.score) if dmi else 0.0,
            "maturity_level": dmi.level_name if dmi else "N/A",
        },
        "prioritized_actions": actions,
        "strategic_summary": _strategic_summary(actions),
    }


def export_selected_assessment(
    backend_results: Mapping[str, Any],
    formats: list[str],
) -> dict[str, Path]:
    """
    Generate selected assessment export formats.

    Supported formats:
        - pdf_score  → 2‑page score summary (PDF)
        - pdf_full   → comprehensive full report (PDF)
        - excel      → Excel workbook
        - json       → JSON data dump
    """
    if not isinstance(backend_results, Mapping):
        raise AssessmentProcessingError("Invalid backend results: expected a mapping.")
    if not isinstance(formats, (list, tuple, set)):
        raise AssessmentProcessingError("Export formats must be a list.")

    formats = {str(fmt).strip().lower() for fmt in formats if fmt}
    if not formats:
        raise AssessmentProcessingError("No export format was selected.")

    assessment_id = str(backend_results.get("assessment_id") or "assessment")
    output_dir = settings.OUTPUT_DIR / assessment_id
    output_dir.mkdir(parents=True, exist_ok=True)

    outputs: dict[str, Path] = {}

    # Normalize data
    aggregation_results = backend_results.get("aggregation") or {}
    gap_results = backend_results.get("gaps") or []
    tpi_results = backend_results.get("tpi") or []
    priority_results = backend_results.get("priorities") or []
    roadmap = backend_results.get("roadmap") or []

    # --------------------------------------------------------------------
    # 1. PDF Score Summary
    # --------------------------------------------------------------------
    if "pdf_score" in formats:
        try:
            from exports.pdf import export_score_summary
            path = export_score_summary(
                aggregation_results=aggregation_results,
                gap_results=gap_results,
                tpi_results=tpi_results,
                priority_results=priority_results,
                output_path=output_dir / f"JESA_DMAT_Score_Summary_{assessment_id}.pdf",
                assessment_id=assessment_id,
            )
            outputs["pdf_score"] = path
        except Exception as e:
            logger.error(f"Error generating PDF score summary: {e}")

    # --------------------------------------------------------------------
    # 2. PDF Full Report
    # --------------------------------------------------------------------
    if "pdf_full" in formats:
        try:
            from exports.pdf import export_full_report
            path = export_full_report(
                aggregation_results=aggregation_results,
                gap_results=gap_results,
                tpi_results=tpi_results,
                priority_results=priority_results,
                roadmap=roadmap,
                output_path=output_dir / f"JESA_DMAT_Full_Report_{assessment_id}.pdf",
                assessment_id=assessment_id,
            )
            outputs["pdf_full"] = path
        except Exception as e:
            logger.error(f"Error generating PDF full report: {e}")

    # --------------------------------------------------------------------
    # 3. Excel Workbook
    # --------------------------------------------------------------------
    if "excel" in formats:
        try:
            from exports.excel import ExcelExporter
            path = ExcelExporter().export_assessment_results(
                aggregation_results=aggregation_results,
                gap_results=gap_results,
                tpi_results=tpi_results,
                priority_results=priority_results,
                roadmap=roadmap,
                output_path=output_dir / f"JESA_DMAT_Workbook_{assessment_id}.xlsx",
                assessment_id=assessment_id,
            )
            outputs["excel"] = path
        except Exception as e:
            logger.error(f"Error generating Excel workbook: {e}")

    # --------------------------------------------------------------------
    # 4. JSON
    # --------------------------------------------------------------------
    if "json" in formats:
        try:
            json_path = output_dir / f"JESA_DMAT_Report_{assessment_id}.json"
            serialized = serialize_backend_results(backend_results)
            json_path.write_text(json.dumps(serialized, indent=2, ensure_ascii=False), encoding="utf-8")
            outputs["json"] = json_path
        except Exception as e:
            logger.error(f"Error generating JSON: {e}")

    return outputs


def serialize_backend_results(
    results: Mapping[str, Any],
) -> dict[str, Any]:
    """Convert backend result objects to JSON-safe data."""

    if not isinstance(results, Mapping):
        raise AssessmentProcessingError("Backend results must be a mapping.")

    aggregation = results.get("aggregation") or {}
    if not isinstance(aggregation, Mapping):
        aggregation = {}

    recommendations_raw = results.get("recommendations") or {}
    if not isinstance(recommendations_raw, Mapping):
        recommendations_raw = {}

    gaps = results.get("gaps") or []
    tpi_results = results.get("tpi") or []
    priorities = results.get("priorities") or []
    roadmap = results.get("roadmap") or []

    return {
        "assessment_id": results.get("assessment_id"),
        "metadata": results.get("metadata") or {},
        "aggregation": {
            "indicators": _serialize_score_map(aggregation.get("indicators") or {}),
            "subdimensions": _serialize_score_map(aggregation.get("subdimensions") or {}),
            "dimensions": _serialize_score_map(aggregation.get("dimensions") or {}),
            "pillars": _serialize_score_map(aggregation.get("pillars") or {}),
            "dmi": _serialize_score(aggregation.get("dmi")),
            "metadata": aggregation.get("metadata") or {},
        },
        "gaps": [gap.to_dict() for gap in gaps if hasattr(gap, "to_dict")],
        "recommendations": {
            str(dim_id): [rec.to_dict() for rec in (recs or []) if hasattr(rec, "to_dict")]
            for dim_id, recs in recommendations_raw.items()
        },
        "tpi": [item.to_dict() for item in tpi_results if hasattr(item, "to_dict")],
        "priorities": [item.to_dict() for item in priorities if hasattr(item, "to_dict")],
        "roadmap": [phase.to_dict() for phase in roadmap if hasattr(phase, "to_dict")],
        "summary": results.get("summary") or {},
    }


def deserialize_backend_results(data: Mapping[str, Any]) -> dict[str, Any]:
    """Restore stored JSON data to backend dataclass objects."""

    aggregation_data = data.get("aggregation", {})
    aggregation = {
        "indicators": _deserialize_score_map(aggregation_data.get("indicators", {})),
        "subdimensions": _deserialize_score_map(aggregation_data.get("subdimensions", {})),
        "dimensions": _deserialize_score_map(aggregation_data.get("dimensions", {})),
        "pillars": _deserialize_score_map(aggregation_data.get("pillars", {})),
        "dmi": _score_from_dict(aggregation_data.get("dmi")),
        "metadata": aggregation_data.get("metadata", {}),
    }
    recommendations = {
        dim_id: [_recommendation_from_dict(rec) for rec in recs]
        for dim_id, recs in (data.get("recommendations") or {}).items()
    }
    return {
        "assessment_id": data.get("assessment_id"),
        "metadata": data.get("metadata", {}),
        "aggregation": aggregation,
        "gaps": [_gap_from_dict(item) for item in data.get("gaps", [])],
        "recommendations": recommendations,
        "tpi": [_tpi_from_dict(item) for item in data.get("tpi", [])],
        "priorities": [_priority_from_dict(item) for item in data.get("priorities", [])],
        "roadmap": [_roadmap_phase_from_dict(item) for item in data.get("roadmap", [])],
        "summary": data.get("summary", {}),
    }


def _merge_metadata(
    metadata: AssessmentMetadata,
    assessment_id: str,
    ui_metadata: Mapping[str, Any],
) -> AssessmentMetadata:
    return replace(
        metadata,
        assessment_id=assessment_id,
        site_name=ui_metadata.get("plant") or metadata.site_name,
        assessment_date=ui_metadata.get("assessment_date") or metadata.assessment_date,
        evaluator_name=ui_metadata.get("assessor_name") or metadata.evaluator_name,
        evaluator_function=ui_metadata.get("assessor_role") or metadata.evaluator_function,
    )


def _target_levels(referentiel: Any) -> dict[str, int]:
    levels = {}
    for dim_id, dimension in referentiel.dimensions.items():
        target = dimension.effective_target_level or dimension.target_level_default or 3
        levels[dim_id] = int(target)
    return levels


def _load_recommendations(gaps: list[GapResult]) -> dict[str, list[RecommendationResult]]:
    if not settings.RECOMMENDATIONS_FILE.exists():
        return {}
    workbook = pd.ExcelFile(settings.RECOMMENDATIONS_FILE)
    recommendations = pd.read_excel(workbook, sheet_name=KnowledgeBaseSheets.RECOMMENDATIONS, header=1)
    trigger_mapping = None
    if KnowledgeBaseSheets.TRIGGER_MAPPING in workbook.sheet_names:
        trigger_mapping = pd.read_excel(workbook, sheet_name=KnowledgeBaseSheets.TRIGGER_MAPPING, header=1)
    return RecommendationEngine(recommendations, trigger_mapping=trigger_mapping).get_all_recommendations(gaps)


def _build_summary(
    aggregation: Mapping[str, Any],
    gaps: list[GapResult],
    tpi: list[TPIResult] | None = None,
    priorities: list[PriorityResult] | None = None,
) -> dict[str, Any]:
    dmi = aggregation.get("dmi")
    return {
        "dmi_score": dmi.score if dmi else None,
        "dmi_level": dmi.level if dmi else None,
        "dmi_level_name": dmi.level_name if dmi else "",
        "critical_gaps": sum(1 for gap in gaps if gap.priority in {"critical", "high"}),
        "priority_dimensions": len(priorities or tpi or [gap for gap in gaps if gap.gap > 0]),
    }


def _target_dmi_percent(gaps: list[GapResult], aggregation: Mapping[str, Any]) -> float:
    if not gaps:
        dmi = aggregation.get("dmi")
        return float(dmi.score) if dmi else 0.0
    return round((sum(gap.target_score for gap in gaps) / len(gaps)) * 20, 1)


def _dashboard_insights(largest_gap: dict[str, Any] | None, strongest: dict[str, Any] | None) -> list[dict[str, str]]:
    insights = []
    if largest_gap and largest_gap["gap"] < 0:
        insights.append(
            {
                "type": "danger",
                "title": "Main Constraint",
                "text": f"{largest_gap['name']} is the largest maturity gap.",
            }
        )
    if strongest:
        insights.append(
            {
                "type": "success",
                "title": "Strength",
                "text": f"{strongest['name']} is the strongest pillar.",
            }
        )
    if not insights:
        insights.append(
            {
                "type": "success",
                "title": "Balanced Profile",
                "text": "No major negative gap is currently detected.",
            }
        )
    return insights


import re

def _priority_label(value) -> str:
    """Normalise n'importe quelle valeur de priorité en label anglais standard."""
    if value is None:
        return "Medium"
    
    # Déballer les tuples/listes (bug observé : ('HAUTE',))
    if isinstance(value, (list, tuple)):
        value = value[0] if len(value) > 0 else ""
    
    raw = str(value).strip().lower()
    
    # Nettoyer les caractères parasites : parenthèses, apostrophes, points, etc.
    raw = re.sub(r"[^a-zàâäçéèêëïîôöùûüÿ\s]", "", raw).strip()
    
    labels = {
        "critical": "Critical",
        "critique": "Critical",
        "high": "High",
        "haute": "High",
        "medium": "Medium",
        "moyenne": "Medium",
        "low": "Low",
        "faible": "Low",
        "very low": "Very Low",
        "tres faible": "Very Low",
        "très faible": "Very Low",
    }
    return labels.get(raw, "Medium")


def _strategic_summary(actions: list[dict[str, Any]]) -> str:
    if not actions:
        return "No prioritized transformation action is currently available."
    top = actions[0]
    return f"Focus first on {top['title']} with a {top['priority_level']} priority."


def _serialize_score_map(values: Mapping[str, Any]) -> dict[str, Any]:
    return {key: _serialize_score(value) for key, value in values.items()}


def _serialize_score(value: Any) -> dict[str, Any] | None:
    if value is None:
        return None
    if isinstance(value, ScoreResult):
        return value.to_dict()
    if isinstance(value, Mapping):
        return dict(value)
    return None


def _deserialize_score_map(values: Mapping[str, Any]) -> dict[str, ScoreResult]:
    return {key: _score_from_dict(value) for key, value in values.items() if value}


def _score_from_dict(data: Mapping[str, Any] | None) -> ScoreResult | None:
    if not data:
        return None
    return ScoreResult(
        entity_id=data.get("entity_id", ""),
        entity_name=data.get("entity_name", ""),
        entity_type=data.get("entity_type", ""),
        score=float(data.get("score", 0)),
        level=int(data.get("level", 0)),
        level_name=data.get("level_name", ""),
        parent_id=data.get("parent_id"),
        children_scores=data.get("children_scores"),
        applicability=data.get("applicability", "Applicable"),
        gap_to_target=data.get("gap_to_target"),
        details=data.get("details"),
    )


def _gap_from_dict(data: Mapping[str, Any]) -> GapResult:
    return GapResult(
        entity_id=data.get("entity_id", ""),
        entity_name=data.get("entity_name", ""),
        entity_type=data.get("entity_type", ""),
        current_score=float(data.get("current_score", 0)),
        target_score=float(data.get("target_score", 0)),
        gap=float(data.get("gap", 0)),
        gap_percent=float(data.get("gap_percent", 0)),
        priority=data.get("priority", "none"),
        details=data.get("details"),
    )


def _tpi_from_dict(data: Mapping[str, Any]) -> TPIResult:
    return TPIResult(
        dimension_id=data.get("dimension_id", ""),
        dimension_name=data.get("dimension_name", ""),
        tpi_score=float(data.get("tpi_score", 0)),
        priority_category=data.get("priority_category", ""),
        gap=float(data.get("gap", 0)),
        business_impact=int(data.get("business_impact", 1)),
        strategic_importance=int(data.get("strategic_importance", 1)),
        expected_roi=int(data.get("expected_roi", 1)),
        implementation_cost=int(data.get("implementation_cost", 1)),
        implementation_difficulty=int(data.get("implementation_difficulty", 1)),
        details=data.get("details"),
    )


def _recommendation_from_dict(data: Mapping[str, Any]) -> RecommendationResult:
    return RecommendationResult(
        recommendation_id=data.get("recommendation_id", ""),
        title=data.get("title", ""),
        description=data.get("description", ""),
        dimension_id=data.get("dimension_id", ""),
        pillar_id=data.get("pillar_id", ""),
        current_maturity_band=data.get("current_maturity_band", ""),
        gap_trigger_code=data.get("gap_trigger_code", ""),
        priority=data.get("priority", "medium"),
        effort=data.get("effort", ""),
        horizon=data.get("horizon", ""),
        detailed_actions=list(data.get("detailed_actions") or []),
        prerequisites=list(data.get("prerequisites") or []),
        dependencies=list(data.get("dependencies") or []),
        completion_evidence=list(data.get("completion_evidence") or []),
        expected_impact=data.get("expected_impact", ""),
        source_reference=data.get("source_reference", ""),
    )


def _priority_from_dict(data: Mapping[str, Any]) -> PriorityResult:
    return PriorityResult(
        dimension_id=data.get("dimension_id", ""),
        dimension_name=data.get("dimension_name", ""),
        current_score=float(data.get("current_score", 0)),
        target_score=float(data.get("target_score", 0)),
        gap=float(data.get("gap", 0)),
        tpi_score=data.get("tpi_score"),
        priority_category=data.get("priority_category", ""),
        recommendations=[_recommendation_from_dict(item) for item in data.get("recommendations", [])],
        details=data.get("details") or {},
    )


def _roadmap_phase_from_dict(data: Mapping[str, Any]) -> RoadmapPhase:
    phase = RoadmapPhase(
        phase_id=data.get("phase_id", ""),
        phase_name=data.get("phase_name", ""),
        horizon=data.get("horizon", ""),
    )
    from engines.decision.roadmap import RoadmapItem

    for item in data.get("items", []):
        phase.items.append(
            RoadmapItem(
                recommendation_id=item.get("recommendation_id", ""),
                title=item.get("title", ""),
                dimension_id=item.get("dimension_id", ""),
                pillar_id=item.get("pillar_id", ""),
                priority=item.get("priority", ""),
                phase=item.get("phase", ""),
                horizon=item.get("horizon", ""),
                effort=item.get("effort", ""),
                tpi_score=item.get("tpi_score"),
                gap=item.get("gap"),
                detailed_actions=list(item.get("detailed_actions") or []),
                prerequisites=list(item.get("prerequisites") or []),
                dependencies=list(item.get("dependencies") or []),
                completion_evidence=list(item.get("completion_evidence") or []),
                expected_impact=item.get("expected_impact", ""),
                source_reference=item.get("source_reference", ""),
            )
        )
    return phase